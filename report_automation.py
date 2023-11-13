#Импорт библиотек
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
import string
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Border,Side


def automatic_file(file_name): #Автоматизация
    #Чтение файла Excel
    exel = pd.read_excel(file_name)
    excel[['ID', 'Контролер', 'Сотрудник', 'Дата оценки', 'Месяц диалога', 'Год диалога', 'Оценка',
           'Комментарий к оценке', 'Месяц диалога СММ', 'Тип оператора', 'Тематика задания СММ', 'Площадка']]
    #Создание сводной таблицы
    table = excel.pivot_table(index='Месяц диалога СММ', columns='Сотрудник', values='Оценка', aggfunc='mean',
                              margins=True, margins_name='Средний балл').round(1)
    #Экспорт сводной таблицы в файл Excel
    table.to_excel('Сводный отчёт.xlsx', sheet_name='Свод', startrow=2)
    #Создание отчета
    wb = load_workbook('Сводный отчёт.xlsx')
    sheet = 'Свод'
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row
    #Построение диаграмм
    wb = load_workbook('Сводный отчёт.xlsx')
    barchart = BarChart()
    data = Reference(sheet,
                     min_col=min_column + 1,
                     max_col=max_column - 1,
                     min_row=min_row,
                     max_row=max_row)
    categories = Reference(sheet,  
                           min_col=min_column,  
                           max_col=min_column,
                           min_row=min_row + 1,
                           max_row=max_row)
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)
    sheet.add_chart(barchart, "A11")
    barchart.title = "Динамика среднего балла"
    barchart.style = 2
    alphabet = list(string.ascii_uppercase)
    excel_alph = alphabet[0:max_column]
    excel_alph
    for i in excel_alph:
        if i != "A":
            sheet[f"{i}{max_row}"].fill = PatternFill(fill_type="solid", start_color='FFBB00', end_color='FFBB00')
    #Форматирование
    thins = Side(border_style="thin", color="000000")
    sheet[f"{excel_alph[0]}{max_row}"].border = Border(top=thins, bottom=thins, left=thins, right=thins)
    sheet[f"{excel_alph[0]}{max_row}"].fill = PatternFill(fill_type="solid", start_color='FFBB00', end_color='FFBB00')
    s = wb.active
    ws.column_dimensions['A'].width = 22
    sheet["A1"] = "Динамика среднего балла по месяцам"
    zagolovok = sheet["A1"]
    zagolovok.font = Font(size=15, color='000000', bold=True, italic=False)
    wb.save('Отчёт контроль качества.xlsx')
    return
automatic_file('download.xlsx')
